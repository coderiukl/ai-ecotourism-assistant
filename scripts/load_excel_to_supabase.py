from __future__ import annotations

import argparse
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATA_TABLES = [
    "destinations",
    "media_assets",
    "services_pricing",
    "operation_hours",
    "activities",
    "faq",
    "seo_keywords",
    "knowledge_base",
    "itineraries",
    "tags",
    "maps_navigation",
    "travel_tips",
    "weather_info",
    "transportation",
    "photo_spots",
    "food_stay",
    "contact_info",
    "events_culture",
    "safety_emergency",
    "chatbot_intents",
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value).strip()


def identifier(value: str, fallback: str) -> str:
    text = re.sub(r"[^a-z0-9_]+", "_", clean(value).lower()).strip("_")
    if not text:
        text = fallback
    if text[0].isdigit():
        text = f"data_{text}"
    return text


def normalize_dsn(dsn: str) -> str:
    dsn = dsn.replace("postgresql+asyncpg://", "postgresql://")
    return dsn.replace("ssl=require", "sslmode=require")


def read_excel(path: Path) -> dict[str, tuple[list[str], list[list[str]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, tuple[list[str], list[list[str]]]] = {}
    for worksheet in workbook.worksheets:
        table = identifier(worksheet.title, "sheet")
        if table not in DATA_TABLES:
            continue

        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            continue

        headers: list[str] = []
        seen: dict[str, int] = {}
        for index, header in enumerate(header_row, start=1):
            name = identifier(clean(header), f"column_{index}")
            seen[name] = seen.get(name, 0) + 1
            if seen[name] > 1:
                name = f"{name}_{seen[name]}"
            headers.append(name)

        values: list[list[str]] = []
        for row in rows:
            record = [clean(row[index] if index < len(row) else "") for index in range(len(headers))]
            if any(record):
                values.append(record)
        result[table] = (headers, values)
    return result


def reset_table(cursor, table: str, headers: list[str]) -> None:
    from psycopg2 import sql

    cursor.execute(sql.SQL("drop table if exists public.{} cascade").format(sql.Identifier(table)))
    cursor.execute(
        sql.SQL("create table public.{} ({})").format(
            sql.Identifier(table),
            sql.SQL(", ").join(sql.SQL("{} text").format(sql.Identifier(header)) for header in headers),
        )
    )


def insert_rows(cursor, table: str, headers: list[str], rows: list[list[str]]) -> None:
    if not rows:
        return

    from psycopg2 import sql

    query = sql.SQL("insert into public.{} ({}) values ({})").format(
        sql.Identifier(table),
        sql.SQL(", ").join(sql.Identifier(header) for header in headers),
        sql.SQL(", ").join(sql.Placeholder() for _ in headers),
    )
    cursor.executemany(query, rows)


def load_to_supabase(database_url: str, excel_path: Path) -> None:
    import psycopg2

    sheets = read_excel(excel_path)
    if not sheets:
        raise SystemExit("No supported sheets found in Excel.")

    with psycopg2.connect(normalize_dsn(database_url)) as conn:
        with conn.cursor() as cursor:
            for table, (headers, rows) in sheets.items():
                reset_table(cursor, table, headers)
                insert_rows(cursor, table, headers, rows)
                print(f"{table}: {len(rows)} rows")
        conn.commit()


def main() -> None:
    parser = argparse.ArgumentParser(description="Load local Excel data into Supabase data tables.")
    parser.add_argument("--excel", default="nui_ba_den_tourism_database.xlsx")
    parser.add_argument("--database-url", default=os.getenv("DATABASE_URL", ""))
    parser.add_argument("--yes", action="store_true", help="Confirm replacing data tables.")
    args = parser.parse_args()

    if not args.yes:
        raise SystemExit("Add --yes to replace Supabase data tables.")
    if not args.database_url:
        raise SystemExit("DATABASE_URL is required.")

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    load_to_supabase(args.database_url, excel_path)


if __name__ == "__main__":
    main()
