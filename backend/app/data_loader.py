import openpyxl

from .config import IMAGES_DIR, XLSX_PATH
from .constants import DEFAULT_ACTIVITIES
from .text_utils import is_valid_image_url, slugify


# =========================================================
# Core Excel helpers
# =========================================================

def load_xlsx():
    """Load workbook from configured XLSX_PATH."""
    try:
        return openpyxl.load_workbook(XLSX_PATH, data_only=True)
    except Exception:
        return None


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_sheet_name(sheet_name):
    return clean_value(sheet_name).strip()


def load_sheet_records(workbook, sheet_name, required_field=None):
    """
    Read one sheet into a list of dictionaries.

    - First row is treated as header.
    - Empty headers are ignored.
    - Empty rows are skipped.
    - If required_field is provided, rows missing that field are skipped.
    """
    if not workbook or sheet_name not in workbook.sheetnames:
        return []

    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [clean_value(header) for header in rows[0]]
    records = []

    for row in rows[1:]:
        record = {}

        for header, value in zip(headers, row):
            if header:
                record[header] = value

        if required_field:
            if clean_value(record.get(required_field)):
                records.append(record)
        else:
            if any(clean_value(value) for value in record.values()):
                records.append(record)

    return records


def get_first(record, keys, default=""):
    for key in keys:
        value = record.get(key)
        if clean_value(value):
            return value
    return default


def record_to_text(title, record):
    parts = [clean_value(title)]

    for key, value in record.items():
        value = clean_value(value)
        if value:
            parts.append(f"{key}: {value}")

    return "\n".join(part for part in parts if clean_value(part))


# =========================================================
# Dynamic sheet loading
# =========================================================

IGNORE_SHEETS = {
    "README",
    "README_v2",
    "template",
    "templates",
    "metadata",
    "schema",
    "guide",
    "guides",
    "operation_hours_guide",
}


def should_ignore_sheet(sheet_name):
    normalized = normalize_sheet_name(sheet_name)
    return normalized in IGNORE_SHEETS or normalized.startswith("_")


def get_main_structured_sheets(workbook):
    """Return all real data sheets from the workbook."""
    if not workbook:
        return set()

    return {
        sheet_name
        for sheet_name in workbook.sheetnames
        if not should_ignore_sheet(sheet_name)
    }


def load_all_sheets(workbook):
    """
    Load every non-system sheet into a dict:

    ALL_SHEETS = {
        "destinations": [row_dict, ...],
        "services_pricing": [row_dict, ...],
        "operation_hours": [row_dict, ...],
        ...
    }
    """
    if not workbook:
        return {}

    all_sheets = {}

    for sheet_name in workbook.sheetnames:
        if should_ignore_sheet(sheet_name):
            continue

        records = load_sheet_records(workbook, sheet_name)
        all_sheets[sheet_name] = records

    return all_sheets


TITLE_KEYS = [
    "title",
    "name",
    "destination_name",
    "service_name",
    "ticket_name",
    "activity_name",
    "question",
    "keyword",
    "primary_keyword",
    "place_name",
    "spot_name",
    "event_name",
    "organization",
    "intent_name",
    "scenario",
    "topic",
    "route",
    "route_or_area",
    "location_name",
    "category",
]

ID_KEYS = [
    "id",
    "dest_id",
    "dest_code",
    "media_id",
    "service_id",
    "pricing_id",
    "activity_id",
    "faq_id",
    "kb_id",
    "operation_id",
    "question_id",
    "tip_id",
    "route_id",
    "spot_id",
    "item_id",
    "contact_id",
    "event_id",
    "safety_id",
    "intent_id",
    "keyword_id",
    "itinerary_id",
    "tag_id",
    "map_id",
    "weather_id",
]

SOURCE_URL_KEYS = [
    "source_url",
    "website",
    "booking_url",
    "google_maps_search_url",
    "image_source_url",
]


def build_record_id(sheet_name, record, index):
    for key in ID_KEYS:
        value = clean_value(record.get(key))
        if value:
            return value

    sheet_specific_id = clean_value(record.get(f"{sheet_name}_id"))
    if sheet_specific_id:
        return sheet_specific_id

    return str(index)


def build_title(record, fallback):
    values = []

    for key in TITLE_KEYS:
        value = clean_value(record.get(key))
        if value:
            values.append(value)

        if len(values) >= 2:
            break

    return " - ".join(values) if values else fallback


def build_keywords(record):
    return " ".join(
        clean_value(value)
        for value in record.values()
        if clean_value(value)
    )


def record_to_knowledge_item(sheet_name, record, index):
    item_id = build_record_id(sheet_name, record, index)
    fallback_title = f"{sheet_name} #{index}"
    title = build_title(record, fallback_title)
    source_url = get_first(record, SOURCE_URL_KEYS)

    return {
        "kb_id": f"{sheet_name}:{item_id}",
        "topic": sheet_name,
        "title": title,
        "content": record_to_text(f"Sheet: {sheet_name}\nTitle: {title}", record),
        "keywords": build_keywords(record),
        "source_url": source_url,
        "sheet_name": sheet_name,
        "row_index": index,
    }


def load_all_sheets_as_knowledge(workbook):
    """
    Convert every row from every non-system sheet into KB items for RAG.
    This is the main function that makes the project expandable:
    add a new sheet to Excel -> it appears in KB_LIST automatically.
    """
    if not workbook:
        return []

    kb_items = []

    for sheet_name in workbook.sheetnames:
        if should_ignore_sheet(sheet_name):
            continue

        records = load_sheet_records(workbook, sheet_name)

        for index, record in enumerate(records, start=1):
            if not any(clean_value(value) for value in record.values()):
                continue

            kb_items.append(record_to_knowledge_item(sheet_name, record, index))

    return kb_items


# =========================================================
# Media assets
# =========================================================

def load_media_assets(workbook):
    records = load_sheet_records(workbook, "media_assets", "media_id")
    media_assets = {}

    for record in records:
        dest_id = clean_value(
            get_first(record, ["dest_id", "dest_code", "place_id"])
        )

        media_type = clean_value(
            get_first(record, ["media_type", "type"])
        ).lower()

        if not dest_id:
            continue

        image_url = get_first(record, ["image_url", "url", "Ảnh"])

        media_assets[dest_id] = {
            "media_id": get_first(record, ["media_id", "id"]),
            "image_url": clean_value(image_url) if is_valid_image_url(image_url) else "",
            "image_caption": get_first(
                record,
                ["caption", "image_caption", "title", "spot_name"],
            ),
            "image_alt": get_first(
                record,
                ["alt_text", "image_alt", "caption", "image_caption"],
            ),
            "image_source_url": get_first(record, ["source_url", "image_source_url"]),
            "media_type": media_type or "image",
        }

    return media_assets


LOCAL_IMAGE_FALLBACK = {
    "TN005": "vuon_vo_nga.png",
    "TN007": "tinh_xa_ngoc_truyen.webp",
    "TN009": "linh_son_phuoc_trung_tu.webp",
    "TN010": "can_cu_dong_kim_quang.webp",
    "TN012": "dong_ba_co_quan_am_tu.png",
    "TN013": "tang_da_nut_doi.png",
    "TN014": "linh_son_tien_thach.webp",
    "TN016": "tuong_phat_nhap_niet_ban.png",
}


def _local_image_path(dest_code):
    fallback = LOCAL_IMAGE_FALLBACK.get(dest_code)

    if not fallback:
        return None

    candidate = IMAGES_DIR / fallback

    return f"/images/{candidate.name}" if candidate.exists() else None


def local_image_url_for(destination, media_asset):
    direct = _local_image_path(clean_value(destination.get("dest_code")))

    if direct:
        return direct

    if not IMAGES_DIR.exists():
        return ""

    candidates = [
        slugify(destination.get("name")),
        slugify(destination.get("dest_code")),
        slugify(media_asset.get("image_caption")),
        slugify(media_asset.get("image_alt")),
    ]

    image_files = [
        path
        for path in IMAGES_DIR.iterdir()
        if path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}
    ]

    for candidate in candidates:
        if not candidate:
            continue

        for image_file in image_files:
            file_slug = slugify(image_file.stem)

            if candidate in file_slug or file_slug in candidate:
                return f"/images/{image_file.name}"

    return f"/images/{image_files[0].name}" if image_files else ""


def apply_media_assets(destinations, media_assets):
    for destination in destinations.values():
        dest_code = clean_value(destination.get("dest_code"))
        media_asset = media_assets.get(dest_code, {})

        external = media_asset.get("image_url")
        image_url = local_image_url_for(destination, media_asset) or external or ""

        destination["image_url"] = image_url
        destination["image_caption"] = media_asset.get("image_caption") or destination.get("name")
        destination["image_alt"] = media_asset.get("image_alt") or destination.get("name")
        destination["image_source_url"] = (
            media_asset.get("image_source_url")
            or destination.get("source_url")
        )

    return destinations


# =========================================================
# Destinations legacy loader for current UI compatibility
# =========================================================

def fallback_destination():
    return {
        "id": 1,
        "dest_code": "TN001",
        "name": "Núi Bà Đen",
        "category": "Tâm linh / Cảnh quan",
        "location": "Tây Ninh, Việt Nam",
        "short_description": (
            "Khám phá biểu tượng du lịch sinh thái, tâm linh và cảnh quan "
            "nổi bật của Tây Ninh."
        ),
        "description_detail": (
            "Núi Bà Đen là điểm đến kết hợp du lịch tâm linh, cảnh quan "
            "và trải nghiệm cáp treo."
        ),
        "highlight": "Tượng Phật Bà; chóp đỉnh 986m; cáp treo",
        "best_time": "Sáng sớm hoặc chiều mát; đẹp nhất mùa khô 12-4",
        "estimated_duration": "1 ngày",
        "travel_type": "Tâm linh / Check-in",
        "difficulty_level": "Dễ",
        "transport": "Cáp treo tuyến Vân Sơn",
        "seo_keyword": "núi bà đen tây ninh",
        "source_url": "",
        "verification_note": "",
        "video_url": "https://example.com/video-nui-ba-den.mp4",
        "activities": DEFAULT_ACTIVITIES,
    }


def load_destinations(workbook):
    records = load_sheet_records(workbook, "destinations", "dest_id")

    if not records:
        return {1: fallback_destination()}

    destinations = {}

    for idx, record in enumerate(records, start=1):
        dest_code = get_first(record, ["dest_id", "dest_code", "Mã"])

        destinations[idx] = {
            "id": idx,
            "dest_code": dest_code,
            "name": get_first(record, ["name", "Tên", "destination_name"]),
            "category": get_first(record, ["category", "Danh mục"]),
            "location": get_first(record, ["location", "Vị trí"]),
            "short_description": get_first(
                record,
                ["description_short", "short_description", "Mô tả ngắn"],
            ),
            "description_detail": get_first(
                record,
                ["description_detail", "detail", "Chi tiết"],
            ),
            "highlight": get_first(record, ["highlight", "Điểm nổi bật"]),
            "best_time": get_first(record, ["best_time", "Thời điểm đẹp"]),
            "estimated_duration": get_first(record, ["estimated_duration", "Thời lượng"]),
            "travel_type": get_first(record, ["travel_type", "Loại hình"]),
            "difficulty_level": get_first(record, ["difficulty_level", "Độ khó"]),
            "transport": get_first(record, ["transport", "Di chuyển"]),
            "seo_keyword": get_first(record, ["seo_keyword", "keywords"]),
            "source_url": get_first(record, ["source_url"]),
            "verification_note": get_first(record, ["verification_note"]),
            "video_url": "https://example.com/video-nui-ba-den.mp4",
            "activities": DEFAULT_ACTIVITIES,
        }

    return destinations or {1: fallback_destination()}


# =========================================================
# Legacy structured loaders
# Giữ lại để code cũ không bị lỗi, nhưng dữ liệu chính nên dùng ALL_SHEETS / KB_LIST.
# =========================================================

def load_faqs(workbook):
    return load_sheet_records(workbook, "faq", "faq_id")


def load_services_pricing(workbook):
    # Không ép service_id, vì sheet mới có thể dùng pricing_id.
    return load_sheet_records(workbook, "services_pricing")


def load_activities(workbook):
    return load_sheet_records(workbook, "activities", "activity_id")


def load_operation_hours(workbook):
    return load_sheet_records(workbook, "operation_hours")


def load_knowledge_base(workbook):
    # RAG dùng toàn bộ sheet, không chỉ sheet knowledge_base.
    return load_all_sheets_as_knowledge(workbook)


# =========================================================
# Load data once at import time
# =========================================================

WORKBOOK = load_xlsx()

MAIN_STRUCTURED_SHEETS = get_main_structured_sheets(WORKBOOK)

ALL_SHEETS = load_all_sheets(WORKBOOK)

MEDIA_ASSETS = load_media_assets(WORKBOOK)

DESTINATIONS = apply_media_assets(
    load_destinations(WORKBOOK),
    MEDIA_ASSETS,
)

# Legacy aliases: giữ để các file đang import biến cũ vẫn chạy.
FAQ_LIST = ALL_SHEETS.get("faq", [])
SERVICES_PRICING = ALL_SHEETS.get("services_pricing", [])
ACTIVITIES = ALL_SHEETS.get("activities", [])
OPERATION_HOURS = ALL_SHEETS.get("operation_hours", [])
KNOWLEDGE_BASE_RAW = ALL_SHEETS.get("knowledge_base", [])

# Main RAG list: tự động bao gồm tất cả sheet không bị ignore.
KB_LIST = load_knowledge_base(WORKBOOK)