from __future__ import annotations

import logging
from functools import lru_cache
from typing import Any

from openpyxl import load_workbook
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except Exception:  # pragma: no cover - optional deploy fallback
    psycopg2 = None
    RealDictCursor = None

from app.core.config import DATABASE_URL, DATA_PATH, IMAGES_DIR

logger = logging.getLogger(__name__)

LOCAL_IMAGE_BY_DEST_ID = {
    "TN001": "tuong_phat_ba_tay_bo_da_son.jpg",
    "TN002": "ton-tuong-bo-tat-di-lac.jpg",
    "TN004": "cum_cot_kinh_bat_nha.jpg",
    "TN005": "vuon_vo_nga.png",
    "TN006": "ga_ba_den.webp",
    "TN007": "tinh_xa_ngoc_truyen.webp",
    "TN008": "tuong_dai_dung_si_nui_ba_den.webp",
    "TN009": "linh_son_phuoc_trung_tu.webp",
    "TN010": "can_cu_dong_kim_quang.webp",
    "TN011": "thap_vang_sanh.jpg",
    "TN012": "dong_ba_co_quan_am_tu.png",
    "TN013": "tang_da_nut_doi.png",
    "TN014": "linh_son_tien_thach.webp",
    "TN015": "dai-hong-chung.jpg",
    "TN016": "tuong_phat_nhap_niet_ban.png",
}

DB_SHEETS = [
    "destinations",
    "media_assets",
    "services_pricing",
    "operation_hours",
    "activities",
    "faq",
    "seo_keywords",
    "knowledge_base",
    "itineraries",
    "tags",
    "maps_navigation",
    "travel_tips",
    "weather_info",
    "transportation",
    "photo_spots",
    "food_stay",
    "contact_info",
    "events_culture",
    "safety_emergency",
    "chatbot_intents",
]


def image_url_for_dest(dest_id: str, fallback_url: str = "") -> str:
    if fallback_url.startswith(("http://", "https://", "/")):
        return fallback_url
    filename = LOCAL_IMAGE_BY_DEST_ID.get(dest_id)
    if filename and (IMAGES_DIR / filename).exists():
        return f"/images/{filename}"
    return ""


def _normalize_postgres_dsn(dsn: str) -> str:
    dsn = dsn.replace("postgresql+asyncpg://", "postgresql://")
    return dsn.replace("ssl=require", "sslmode=require")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value).strip()


def _row_to_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, str]:
    record: dict[str, str] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        value = clean(row[index] if index < len(row) else "")
        if value:
            record[header] = value
    return record


def _load_supabase_rows() -> dict[str, list[dict[str, str]]]:
    if not DATABASE_URL or psycopg2 is None or RealDictCursor is None:
        return {}

    sheets: dict[str, list[dict[str, str]]] = {}
    try:
        with psycopg2.connect(_normalize_postgres_dsn(DATABASE_URL), cursor_factory=RealDictCursor) as conn:
            with conn.cursor() as cursor:
                for table in DB_SHEETS:
                    try:
                        cursor.execute(f"select * from {table}")
                    except Exception as exc:
                        conn.rollback()
                        logger.warning("Supabase table skipped: %s (%s)", table, exc)
                        continue
                    rows = cursor.fetchall()
                    sheets[table] = [
                        {key: clean(value) for key, value in dict(row).items() if clean(value)}
                        for row in rows
                    ]
    except Exception as exc:
        logger.warning("Supabase data load skipped: %s", exc)
        return {}

    return sheets


@lru_cache(maxsize=1)
def load_workbook_rows() -> dict[str, list[dict[str, str]]]:
    if DATABASE_URL:
        return _load_supabase_rows()

    if not DATA_PATH.exists():
        logger.warning("Excel file not found: %s", DATA_PATH)
        return {}

    workbook = load_workbook(DATA_PATH, read_only=True, data_only=True)
    sheets: dict[str, list[dict[str, str]]] = {}
    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            continue
        headers = [clean(cell).lower().strip() for cell in header_row]
        records = [_row_to_dict(headers, row) for row in rows]
        sheets[worksheet.title] = [record for record in records if any(record.values())]
    return sheets


@lru_cache(maxsize=1)
def destinations() -> dict[int, dict[str, Any]]:
    sheets = load_workbook_rows()
    media_by_dest: dict[str, dict[str, str]] = {}
    for media in sheets.get("media_assets", []):
        media_by_dest.setdefault(media.get("dest_id", ""), media)

    result: dict[int, dict[str, Any]] = {}
    for index, record in enumerate(sheets.get("destinations", []), start=1):
        dest_id = record.get("dest_id", "")
        media = media_by_dest.get(dest_id, {})
        result[index] = {
            "id": index,
            "dest_code": dest_id,
            "name": record.get("name", "Điểm đến Núi Bà Đen"),
            "category": record.get("category", ""),
            "location": record.get("location", ""),
            "short_description": record.get("description_short", ""),
            "description_detail": record.get("description_detail", ""),
            "highlight": record.get("highlight", ""),
            "best_time": record.get("best_time", ""),
            "estimated_duration": record.get("estimated_duration", ""),
            "travel_type": record.get("travel_type", ""),
            "difficulty_level": record.get("difficulty_level", ""),
            "transport": record.get("transport", ""),
            "seo_keyword": record.get("seo_keyword", ""),
            "source_url": record.get("source_url", ""),
            "image_url": image_url_for_dest(dest_id, media.get("image_url", "")),
            "image_caption": media.get("caption", ""),
            "image_alt": media.get("alt_text", ""),
        }
    return result


def _title_for_record(sheet: str, record: dict[str, str]) -> str:
    for key in ("name", "title", "question", "service_name", "activity_name", "place_name", "spot_name", "event_name", "organization", "keyword"):
        if record.get(key):
            return record[key]
    return sheet


def _record_text(sheet: str, record: dict[str, str]) -> str:
    labels = [f"sheet: {sheet}", f"title: {_title_for_record(sheet, record)}"]
    labels.extend(f"{key}: {value}" for key, value in record.items() if value)
    return "\n".join(labels)


@lru_cache(maxsize=1)
def rag_documents() -> list[dict[str, Any]]:
    documents: list[dict[str, Any]] = []
    for sheet, records in load_workbook_rows().items():
        for row_index, record in enumerate(records, start=2):
            doc_id = f"{sheet}:{row_index}"
            dest_code = record.get("dest_id") or record.get("place_id") or ""
            documents.append({
                "id": doc_id,
                "sheet": sheet,
                "type": sheet,
                "title": _title_for_record(sheet, record),
                "dest_code": dest_code,
                "source_url": record.get("source_url", ""),
                "text": _record_text(sheet, record),
            })
    return documents


def stats() -> dict[str, Any]:
    sheets = load_workbook_rows()
    return {
        "data_source": "supabase" if DATABASE_URL else "excel",
        "data_path": str(DATA_PATH),
        "data_exists": DATA_PATH.exists(),
        "sheets": {name: len(rows) for name, rows in sheets.items()},
        "destinations": len(destinations()),
        "documents": len(rag_documents()),
    }
