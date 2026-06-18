from __future__ import annotations

import json
import logging
import asyncio
from typing import Any
from functools import lru_cache
from openpyxl import load_workbook
from app.core.config import DATABASE_URL, DATA_PATH

try:
    import asyncpg
except Exception:
    asyncpg = None

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except Exception:
    psycopg2 = None
    RealDictCursor = None


logger = logging.getLogger(__name__)
_pool: Any | None = None

DB_SHEETS = [
    "destinations",
    "media_assets",
    "services_pricing",
    "operation_hours",
    "activities",
    "faq",
    "seo_keywords",
    "knowledge_base",
    "itineraries",
    "tags",
    "maps_navigation",
    "travel_tips",
    "weather_info",
    "transportation",
    "photo_spots",
    "food_stay",
    "contact_info",
    "events_culture",
    "safety_emergency",
    "chatbot_intents",
]

def _normalize_sync_dsn() -> str:
    dsn = DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://")
    return dsn.replace("ssl=require", "sslmode=require")

async def connect() -> None:
    global _pool

    if not DATABASE_URL or _pool is not None:
        return

    if asyncpg is None:
        logger.warning("Postgres disabled: asyncpg is not installed")
        return

    try:
        dsn, kwargs = _normalize_sync_dsn()
        _pool = await asyncio.wait_for(
            asyncpg.create_pool(
                dsn=dsn,
                min_size=0,
                max_size=2,
                timeout=5,
                command_timeout=10,
                **kwargs,
            ),
            timeout=8,
        )
        await asyncio.wait_for(migrate(), timeout=10)
        logger.info("Postgres connected")
    except Exception as exc:
        _pool = None
        logger.warning("Postgres disabled: %s", exc)

async def close() -> None:
    global _pool
    if _pool is not None:
        await _pool.close()
        _pool = None


async def migrate() -> None:
    if _pool is None:
        return

    async with _pool.acquire() as conn:
        await conn.execute(
            """
            create table if not exists chat_sessions (
                session_id text primary key,
                destination_id integer,
                created_at timestamptz not null default now(),
                updated_at timestamptz not null default now()
            );

            create table if not exists chat_messages (
                id bigserial primary key,
                session_id text not null,
                role text not null check (role in ('user','assistant')),
                content text not null,
                metadata jsonb not null default '{}'::jsonb,
                created_at timestamptz not null default now()
            );
            """
        )


async def save_chat(session_id: str, destination_id: int | None, user_message: str, assistant_message: str, metadata: dict[str, Any]) -> None:
    if _pool is None:
        return

    async with _pool.acquire() as conn:
        await conn.execute(
            """
            insert into chat_sessions(session_id, destination_id, updated_at)
            values($1, $2, now())
            on conflict(session_id)
            do update set destination_id = excluded.destination_id, updated_at = now()
            """,
            session_id,
            destination_id,
        )

        await conn.execute(
            "insert into chat_messages(session_id, role, content, metadata) values($1, 'user', $2, '{}'::jsonb)",
            session_id,
            user_message,
        )

        await conn.execute(
            "insert into chat_messages(session_id, role, content, metadata) values($1, 'assistant', $2, $3::jsonb)",
            session_id,
            assistant_message,
            json.dumps(metadata, ensure_ascii=False),
        )


def enabled() -> bool:
    return _pool is not None

def clean(value: Any) -> str:
    if value is None:
        return ""

    if hasattr(value, "isoformat"):
        return value.isoformat()

    return str(value).strip()

def image_url_for_dest(fallback_url: str = "") -> str:
    return fallback_url.strip()

def _row_to_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, str]:
    record: dict[str, str] = {}

    for index, header in enumerate(headers):
        if not header:
            continue

        value = clean(row[index] if index < len(row) else "")

        if value:
            record[header] = value

    return record

def _load_supabase_rows() -> dict[str, list[dict[str, str]]]:
    if not DATABASE_URL or psycopg2 is None or RealDictCursor is None:
        return {}

    sheets: dict[str, list[dict[str, str]]] = {}

    try:
        with psycopg2.connect(
            _normalize_sync_dsn(),
            cursor_factory=RealDictCursor,
        ) as conn:
            with conn.cursor() as cursor:
                for table in DB_SHEETS:
                    try:
                        cursor.execute(f"select * from {table}")
                    except Exception as exc:
                        conn.rollback()
                        logger.warning("Supabase table skipped: %s (%s)", table, exc)
                        continue

                    rows = cursor.fetchall()
                    sheets[table] = [
                        {
                            key: clean(value)
                            for key, value in dict(row).items()
                            if clean(value)
                        }
                        for row in rows
                    ]
    except Exception as exc:
        logger.warning("Supabase data load skipped: %s", exc)
        return {}

    return sheets


@lru_cache(maxsize=1)
def load_rows() -> dict[str, list[dict[str, str]]]:
    if DATABASE_URL:
        return _load_supabase_rows()

    if not DATA_PATH.exists():
        logger.warning("Excel file not found: %s", DATA_PATH)
        return {}

    workbook = load_workbook(DATA_PATH, read_only=True, data_only=True)
    sheets: dict[str, list[dict[str, str]]] = {}

    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)

        if not header_row:
            continue

        headers = [clean(cell).lower().strip() for cell in header_row]
        records = [_row_to_dict(headers, row) for row in rows]
        sheets[worksheet.title] = [record for record in records if any(record.values())]

    return sheets


@lru_cache(maxsize=1)
def destinations() -> dict[int, dict[str, Any]]:
    sheets = load_rows()

    media_by_dest: dict[str, dict[str, str]] = {}
    for media in sheets.get("media_assets", []):
        media_by_dest.setdefault(media.get("dest_id", ""), media)

    result: dict[int, dict[str, Any]] = {}

    for index, record in enumerate(sheets.get("destinations", []), start=1):
        dest_id = record.get("dest_id", "")
        media = media_by_dest.get(dest_id, {})

        result[index] = {
            "id": index,
            "dest_code": dest_id,
            "name": record.get("name", "Điểm đến Núi Bà Đen"),
            "category": record.get("category", ""),
            "location": record.get("location", ""),
            "short_description": record.get("description_short", ""),
            "description_detail": record.get("description_detail", ""),
            "highlight": record.get("highlight", ""),
            "best_time": record.get("best_time", ""),
            "estimated_duration": record.get("estimated_duration", ""),
            "travel_type": record.get("travel_type", ""),
            "difficulty_level": record.get("difficulty_level", ""),
            "transport": record.get("transport", ""),
            "seo_keyword": record.get("seo_keyword", ""),
            "source_url": record.get("source_url", ""),
            "image_url": image_url_for_dest(media.get("image_url", "")),
            "image_caption": media.get("caption", ""),
            "image_alt": media.get("alt_text", ""),
        }

    return result


def _title_for_record(sheet: str, record: dict[str, str]) -> str:
    for key in (
        "name",
        "title",
        "question",
        "service_name",
        "activity_name",
        "place_name",
        "spot_name",
        "event_name",
        "organization",
        "keyword",
    ):
        if record.get(key):
            return record[key]

    return sheet


def _record_text(sheet: str, record: dict[str, str]) -> str:
    labels = [f"sheet: {sheet}", f"title: {_title_for_record(sheet, record)}"]
    labels.extend(f"{key}: {value}" for key, value in record.items() if value)
    return "\n".join(labels)


@lru_cache(maxsize=1)
def rag_documents() -> list[dict[str, Any]]:
    documents: list[dict[str, Any]] = []

    for sheet, records in load_rows().items():
        for row_index, record in enumerate(records, start=2):
            documents.append(
                {
                    "id": f"{sheet}:{row_index}",
                    "sheet": sheet,
                    "type": sheet,
                    "title": _title_for_record(sheet, record),
                    "dest_code": record.get("dest_id") or record.get("place_id") or "",
                    "source_url": record.get("source_url", ""),
                    "text": _record_text(sheet, record),
                }
            )

    return documents


def data_stats() -> dict[str, Any]:
    sheets = load_rows()

    return {
        "data_source": "supabase" if DATABASE_URL else "excel",
        "data_path": str(DATA_PATH),
        "data_exists": DATA_PATH.exists(),
        "sheets": {name: len(rows) for name, rows in sheets.items()},
        "destinations": len(destinations()),
        "documents": len(rag_documents()),
    }